import csv
from io import BytesIO
from pathlib import Path
from openpyxl.reader.excel import load_workbook
from sqlalchemy import select, and_
from sqlalchemy import or_, func
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from sqlalchemy.ext.asyncio import AsyncSession
# from sqlalchemy.sql.functions import current_user

from app.core.security import async_get_db
from app.models.auth import User
from app.schemas.entry_schemas import MedicineCreate, MedicineRead, MedicineUpdate
from app.services.medicine_entry import (
    create_medicine, get_medicine, update_medicine, delete_medicine
)
from app.models.entry_models import Medicine
from app.core.security import get_current_user, require_roles


router = APIRouter(prefix="/entry_medicines", tags=["entry_medicines"])

# @router.post("/", response_model=MedicineRead)
# async def create_medicine_endpoint(
#     medicine: MedicineCreate,
#     db: AsyncSession = Depends(get_db)
# ):
#     return await create_medicine(db, medicine)

# @router.post("/", response_model=MedicineRead)
# async def create_medicine_endpoint(
#     medicine: MedicineCreate,
#     db: AsyncSession = Depends(get_db),
#     pharmacist: User = Depends(role_required("pharmacist")),
# ):
#       # DEBUG - check what pharmacist has
#     print(f"pharmacist.id         = {pharmacist.id}")
#     print(f"pharmacist.username   = {pharmacist.username}")
#     print(f"pharmacist.hospital_id = {pharmacist.hospital_id}")
#     print(f"pharmacist.branch_id   = {pharmacist.branch_id}")

#     medicine_data = medicine.model_dump()

#     medicine_data["hospital_id"] = pharmacist.hospital_id
#     medicine_data["branch_id"] = pharmacist.branch_id

#     print(f"medicine_data = {medicine_data}")  # check final dict

#     return await create_medicine(db, medicine_data)

@router.post("/", response_model=MedicineRead)
async def create_medicine_endpoint(
    medicine: MedicineCreate,
    db: AsyncSession = Depends(async_get_db),
    pharmacist: User = Depends(require_roles("pharmacist","superadmin")),
):
    print(f"[DEBUG] Full user dict: {pharmacist.__dict__}")
    # Block unassigned pharmacists FIRST
    if not pharmacist.hospital_id or not pharmacist.current_branch_id:
        print(pharmacist.hospital_id, pharmacist.current_branch_id)
        raise HTTPException(
            status_code=400,
            detail="Your account is not assigned to a hospital or branch. Contact your admin."
        )

    medicine_data = medicine.model_dump()
    medicine_data["hospital_id"] = pharmacist.hospital_id
    medicine_data["branch_id"] = pharmacist.current_branch_id

    return await create_medicine(db, medicine_data)



# @router.get("/search", response_model=list[MedicineRead])
# async def search_medicines(
#         q: str | None = Query(None, description="A-Z OR item_name/strength/brand_name/company search"),
#         db: AsyncSession = Depends(get_db)
# ):
#     query = select(Medicine)

#     if q:
#         search_term = q.strip().lower()


#         if len(search_term) == 1 and search_term.isalpha():
#             query = query.where(
#                 or_(
#                     func.lower(Medicine.item_name).like(f"{search_term}%"),
#                     func.lower(Medicine.strength).like(f"{search_term}%"),
#                     func.lower(Medicine.brand_name).like(f"{search_term}%"),
#                     func.lower(Medicine.company).like(f"{search_term}%")
#                 )
#             )

#         else:  
#             query = query.where(
#                 or_(
#                     func.lower(Medicine.item_name).like(f"{search_term}%"),
#                     func.lower(Medicine.strength).like(f"{search_term}%"),
#                     func.lower(Medicine.brand_name).like(f"{search_term}%"),
#                     func.lower(Medicine.company).like(f"{search_term}%")
#                 )
#             )

#     query = query.order_by(Medicine.item_name.asc())
#     result = await db.execute(query)
#     return result.scalars().all()

# @router.get("/search", response_model=list[MedicineRead])
# async def search_medicines(
#     # hospital_id: int = Query(...),
#     # branch_id: int = Query(...),
#     q: str | None = Query(None),
#     db: AsyncSession = Depends(get_db)
# ):

#     # query = select(Medicine).where(
#     #     Medicine.hospital_id == hospital_id,
#     #     Medicine.branch_id == branch_id
#     # )

#     if q:
#         term = q.lower()

#         query = query.where(
#             or_(
#                 func.lower(Medicine.item_name).like(f"{term}%"),
#                 func.lower(Medicine.brand_name).like(f"{term}%"),
#                 func.lower(Medicine.company).like(f"{term}%"),
#                 func.lower(Medicine.strength).like(f"{term}%"),
#             )
#         )

#     query = query.order_by(Medicine.item_name.asc())

#     result = await db.execute(query)
#     return result.scalars().all()


# @router.get("/search", response_model=list[MedicineRead])
# async def search_medicines(
#     q: str | None = Query(None),
#     db: AsyncSession = Depends(get_db)
# ):
#
#     query = select(Medicine)   # initialize query
#
#     if q:
#         term = q.lower()
#
#         query = query.where(
#             or_(
#                 func.lower(Medicine.item_name).like(f"{term}%"),
#                 func.lower(Medicine.brand_name).like(f"{term}%"),
#                 func.lower(Medicine.company).like(f"{term}%"),
#                 func.lower(Medicine.strength).like(f"{term}%"),
#             )
#         )
#
#     query = query.order_by(Medicine.item_name.asc())
#
#     result = await db.execute(query)
#     return result.scalars().all()
#
# @router.get("/search", response_model=list[MedicineRead])
# async def search_medicines(
#     q: str | None = Query(None),
#     db: AsyncSession = Depends(get_db),
#     pharmacist: User = Depends(role_required(["PHARMACIST"]))
# ):
#
#     query = select(Medicine).where(
#         Medicine.hospital_id == pharmacist.hospital_id,
#         Medicine.branch_id == pharmacist.branch_id
#     )
#
#     if q:
#         term = q.lower()
#
#         query = query.where(
#             or_(
#                 func.lower(Medicine.item_name).like(f"{term}%"),
#                 func.lower(Medicine.brand_name).like(f"{term}%"),
#                 func.lower(Medicine.company).like(f"{term}%"),
#                 func.lower(Medicine.strength).like(f"{term}%"),
#             )
#         )
#
#     query = query.order_by(Medicine.item_name.asc())
#
#     result = await db.execute(query)
#
#     return result.scalars().all()
@router.get("/search", response_model=list[MedicineRead])
async def search_medicines(
    q: str | None = Query(None),
    db: AsyncSession = Depends(async_get_db),
        pharmacist: User = Depends(
            require_roles("pharmacist", "superadmin")
        )
):

    query = select(Medicine).where(
        Medicine.hospital_id == pharmacist.hospital_id,
        Medicine.branch_id == pharmacist.current_branch_id,

        Medicine.item_name.is_not(None),
        Medicine.hsn_code.is_not(None),
        Medicine.category.is_not(None),
        Medicine.strength.is_not(None),
        Medicine.dosage_form.is_not(None),
        Medicine.price.is_not(None),
    )

    if q:
        term = q.lower().strip()

        query = query.where(
            or_(
                func.lower(Medicine.item_name).contains(term),
                func.lower(Medicine.brand_name).contains(term),
                func.lower(Medicine.company).contains(term),
                func.lower(Medicine.strength).contains(term),
            )
        )

    query = query.order_by(Medicine.item_name.asc())

    result = await db.execute(query)

    return result.scalars().all()


# @router.get("/all", response_model=list[MedicineRead])
# async def get_all_medicines_across_branches(
#     db: AsyncSession = Depends(async_get_db),
#         pharmacist: User = Depends(require_roles(["PHARMACIST", "superadmin"])),
# ):
#     """
#     Get all medicines across all branches for the logged-in user's hospital
#     """
#     query = (
#         select(Medicine)
#         .where(
#             and_(
#                 Medicine.hospital_id == current_user.hospital_id,
#                 Medicine.item_name.isnot(None),
#                 Medicine.hsn_code.isnot(None),
#                 Medicine.category.isnot(None),
#                 Medicine.strength.isnot(None),
#                 Medicine.dosage_form.isnot(None),
#                 Medicine.price.isnot(None),
#             )
#         )
#         .order_by(Medicine.item_name.asc())
#     )
#
#     result = await db.execute(query)
#     return result.scalars().all()
@router.get("/all", response_model=list[MedicineRead])
async def get_all_medicines_across_branches(
    db: AsyncSession = Depends(async_get_db),
    pharmacist: User = Depends(
        require_roles("pharmacist", "superadmin")
    ),
):
    query = (
        select(Medicine)
        .where(
            and_(
                Medicine.hospital_id == pharmacist.hospital_id,
                Medicine.item_name.isnot(None),
                Medicine.hsn_code.isnot(None),
                Medicine.category.isnot(None),
                Medicine.strength.isnot(None),
                Medicine.dosage_form.isnot(None),
                Medicine.price.isnot(None),
            )
        )
        .order_by(Medicine.item_name.asc())
    )

    result = await db.execute(query)
    return result.scalars().all()

# @router.get("/all/search", response_model=list[MedicineRead])
# async def get_all_medicines_search(
#     q: str | None = Query(None),
#     branch_id: int | None = Query(default=None),
#     db: AsyncSession = Depends(async_get_db),
#     pharmacist: User = Depends(require_roles(["PHARMACIST","superadmin"])),
# ):
#     """
#     Get medicines across branches with optional search + branch filter
#     """
#
#     query = (
#         select(Medicine)
#         .where(
#             and_(
#                 Medicine.hospital_id == current_user.hospital_id,
#                 Medicine.item_name.isnot(None),
#                 Medicine.hsn_code.isnot(None),
#                 Medicine.category.isnot(None),
#                 Medicine.strength.isnot(None),
#                 Medicine.dosage_form.isnot(None),
#                 Medicine.price.isnot(None),
#             )
#         )
#         .order_by(Medicine.item_name.asc())
#     )
#     if branch_id:
#         query = query.where(Medicine.branch_id == branch_id)
#
#     if q:
#         term = q.lower()
#         query = query.where(
#             or_(
#                 func.lower(Medicine.item_name).like(f"{term}%"),
#                 func.lower(Medicine.brand_name).like(f"{term}%"),
#                 func.lower(Medicine.company).like(f"{term}%"),
#                 func.lower(Medicine.strength).like(f"{term}%"),
#             )
#         )
#
#     query = query.order_by(Medicine.item_name.asc())
#
#     result = await db.execute(query)
#     return result.scalars().all()
@router.get("/all/search", response_model=list[MedicineRead])
async def get_all_medicines_search(
    q: str | None = Query(None),
    branch_id: int | None = Query(default=None),
    db: AsyncSession = Depends(async_get_db),
    pharmacist: User = Depends(
        require_roles("pharmacist", "superadmin")
    ),
):
    query = (
        select(Medicine)
        .where(
            and_(
                Medicine.hospital_id == pharmacist.hospital_id,
                Medicine.item_name.isnot(None),
                Medicine.hsn_code.isnot(None),
                Medicine.category.isnot(None),
                Medicine.strength.isnot(None),
                Medicine.dosage_form.isnot(None),
                Medicine.price.isnot(None),
            )
        )
    )

    if branch_id:
        query = query.where(Medicine.branch_id == branch_id)

    if q:
        term = q.lower().strip()
        query = query.where(
            or_(
                func.lower(Medicine.item_name).contains(term),
                func.lower(Medicine.brand_name).contains(term),
                func.lower(Medicine.company).contains(term),
                func.lower(Medicine.strength).contains(term),
            )
        )

    query = query.order_by(Medicine.item_name.asc())

    result = await db.execute(query)
    return result.scalars().all()


@router.get("/{medicine_id}", response_model=MedicineRead)
async def get_medicine_endpoint(
    medicine_id: int,
    hospital_id: int,
    branch_id: int,
    db: AsyncSession = Depends(async_get_db),
    pharmacist: User = Depends(require_roles(["PHARMACIST", "superadmin"])),
):
    """
    Get a single medicine by ID, scoped to hospital and branch
    """
    result = await db.execute(
        select(Medicine).where(
            and_(
                Medicine.id == medicine_id,
                Medicine.hospital_id == hospital_id,
                Medicine.branch_id == branch_id,
            )
        )
    )
    medicine = result.scalar_one_or_none()

    if not medicine:
        raise HTTPException(status_code=404, detail="Medicine not found")

    return medicine

# @router.get("/", response_model=List[BatchRead])
# async def list_batches(
#     skip: int = Query(0, ge=0),
#     limit: int = Query(100, le=100),
#     search: Optional[str] = Query(None, description="Search"),
#     db: AsyncSession = Depends(get_db)
# ):
#     batches = await get_batches(db, skip=skip, limit=limit, search=search)
#     return batches


# @router.put("/{medicine_id}", response_model=MedicineRead)
# async def update_medicine_endpoint(
#     medicine_id: int,
#     medicine_update: MedicineUpdate,
#     hospital_id: int,
#     branch_id: int,
#     db: AsyncSession = Depends(async_get_db),
# ):
#     updated = await update_medicine(
#         db,
#         medicine_id,
#         medicine_update,
#         hospital_id,
#         branch_id,
#     )
#
#     if not updated:
#         raise HTTPException(404, "Medicine not found")
#
#     return updated
@router.put("/{medicine_id}", response_model=MedicineRead)
async def update_medicine_endpoint(
    medicine_id: int,
    medicine_update: MedicineUpdate,
    hospital_id: int,
    branch_id: int,
    db: AsyncSession = Depends(async_get_db),
    pharmacist: User = Depends(
        require_roles("pharmacist", "superadmin")
    ),
):
    updated = await update_medicine(
        db,
        medicine_id,
        medicine_update,
        pharmacist.hospital_id,
        pharmacist.current_branch_id,
    )

    if not updated:
        raise HTTPException(status_code=404, detail="Medicine not found")

    return updated


# @router.delete("/{medicine_id}")
# async def delete_medicine_endpoint(
#     medicine_id: int,
#     hospital_id: int,
#     branch_id: int,
#     db: AsyncSession = Depends(async_get_db),
# ):
#     deleted = await delete_medicine(
#         db,
#         medicine_id,
#         hospital_id,
#         branch_id,
#     )
#
#     if not deleted:
#         raise HTTPException(404, "Medicine not found")
#
#     return {"message": "Medicine deleted successfully"}
@router.delete("/{medicine_id}")
async def delete_medicine_endpoint(
    medicine_id: int,
    hospital_id: int,
    branch_id: int,
    db: AsyncSession = Depends(async_get_db),
    pharmacist: User = Depends(
        require_roles("pharmacist", "superadmin")
    ),
):
    deleted = await delete_medicine(
        db,
        medicine_id,
        hospital_id,
        branch_id,
    )

    if not deleted:
        raise HTTPException(status_code=404, detail="Medicine not found")

    return {"message": "Medicine deleted successfully"}


UPLOAD_DIR = Path("uploads")
UPLOAD_DIR.mkdir(exist_ok=True)


@router.post("/upload", response_model=dict)
async def upload_document_or_excel(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(async_get_db),
        pharmacist: User = Depends(
            require_roles("pharmacist", "superadmin")
        ),
):
    """
    Upload documents or import medicines from CSV/XLSX
    Medicines are isolated per hospital + branch
    """
    if not pharmacist.hospital_id or not pharmacist.current_branch_id:
        raise HTTPException(
            status_code=400,
            detail="Your account is not assigned to a hospital or branch."
        )

    hospital_id = pharmacist.hospital_id
    branch_id = pharmacist.current_branch_id

    allowed_types = {
        "application/pdf",
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "text/csv",
        "image/jpeg",
        "image/png",
    }

    if file.content_type not in allowed_types:
        raise HTTPException(
            status_code=400,
            detail="Unsupported file type.",
        )

    file_path = UPLOAD_DIR / file.filename

    try:
        contents = await file.read()

        with open(file_path, "wb") as f:
            f.write(contents)

       
        if file.content_type not in [
            "text/csv",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ]:
            return {
                "message": "Document uploaded (no import)",
                "filename": file.filename,
                "path": str(file_path),
            }

        created = 0
        updated = 0
        errors = []

    
        if file.content_type == "text/csv":

            stream = BytesIO(contents)
            reader = csv.reader(stream.read().decode("utf-8").splitlines())

            header = next(reader, None)
            if not header:
                raise HTTPException(400, "CSV is empty")

            try:
                item_name_idx = header.index("item_name")
                strength_idx = header.index("strength")
                brand_name_idx = header.index("brand_name")
                company_idx = header.index("company")
            except ValueError as e:
                raise HTTPException(400, f"Missing column: {e}")

            for row_idx, row in enumerate(reader, start=2):
                try:
                    item_name = str(row[item_name_idx] or "").strip()
                    strength = str(row[strength_idx] or "").strip()
                    brand_name = str(row[brand_name_idx] or "").strip()
                    company = str(row[company_idx] or "").strip()

                    if not item_name or not strength:
                        raise ValueError("Missing item_name/strength")

                    stmt = select(Medicine).where(
                        Medicine.item_name == item_name,
                        Medicine.strength == strength,
                        Medicine.hospital_id == hospital_id,
                        Medicine.branch_id == branch_id,
                    )

                    result = await db.execute(stmt)
                    existing = result.scalar_one_or_none()

                    if existing:
                        existing.brand_name = brand_name
                        existing.company = company
                        updated += 1
                    else:
                        new_med = Medicine(
                            item_name=item_name,
                            strength=strength,
                            brand_name=brand_name,
                            company=company,
                            hospital_id=hospital_id,
                            branch_id=branch_id,
                        )
                        db.add(new_med)
                        created += 1

                except Exception as e:
                    errors.append(f"Row {row_idx}: {str(e)}")

        else:
            wb = load_workbook(BytesIO(contents), read_only=True)
            ws = wb.active

            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                raise HTTPException(400, "Excel empty")

            header = list(rows[0])

            try:
                item_name_idx = header.index("item_name")
                strength_idx = header.index("strength")
                brand_name_idx = header.index("brand_name")
                company_idx = header.index("company")
            except ValueError as e:
                raise HTTPException(400, f"Missing column: {e}")

            for row_idx, row in enumerate(rows[1:], start=2):
                try:
                    item_name = str(row[item_name_idx] or "").strip()
                    strength = str(row[strength_idx] or "").strip()
                    brand_name = str(row[brand_name_idx] or "").strip()
                    company = str(row[company_idx] or "").strip()

                    if not item_name or not strength:
                        raise ValueError("Missing item_name/strength")

                    stmt = select(Medicine).where(
                        Medicine.item_name == item_name,
                        Medicine.strength == strength,
                        Medicine.hospital_id == hospital_id,
                        Medicine.branch_id == branch_id,
                    )

                    result = await db.execute(stmt)
                    existing = result.scalar_one_or_none()

                    if existing:
                        existing.brand_name = brand_name
                        existing.company = company
                        updated += 1
                    else:
                        db.add(
                            Medicine(
                                item_name=item_name,
                                strength=strength,
                                brand_name=brand_name,
                                company=company,
                                hospital_id=hospital_id,
                                branch_id=branch_id,
                            )
                        )
                        created += 1

                except Exception as e:
                    errors.append(f"Row {row_idx}: {str(e)}")

        
        await db.commit()

        return {
            "message": "Upload processed successfully",
            "filename": file.filename,
            "created": created,
            "updated": updated,
            "errors": errors,
        }

    except Exception as e:
        await db.rollback()
        raise HTTPException(
            status_code=500,
            detail=f"Failed to process file: {str(e)}"
        )